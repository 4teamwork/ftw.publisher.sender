from collective.easyform.api import get_context
from ftw.publisher.sender.interfaces import IConfig
from ftw.publisher.sender.utils import sendRequestToRealm
from ftw.publisher.sender.workflows.interfaces import IPublisherContextState
from StringIO import StringIO
from urllib2 import HTTPError
from zope.component import getMultiAdapter
from zope.component.hooks import getSite
import os


def combine_excel(response, local_excel, remote_excel, has_title_row=False):
    from openpyxl import load_workbook

    local_tmp_file = StringIO(local_excel)
    local_wb = load_workbook(local_tmp_file)
    local_ws = local_wb.active

    if remote_excel:
        remote_tmp_file = StringIO(remote_excel.read())
        remote_wb = load_workbook(remote_tmp_file)
        remote_ws = remote_wb.active
        for index, row in enumerate(remote_ws.rows):
            if has_title_row and index == 0:
                # Do not add a second title row
                continue
            local_ws.append(map(lambda cell: cell.value, row))

    output = StringIO()
    local_wb.save(output)
    result = output.getvalue()
    output.close()

    response.setHeader("Content-Disposition", 'attachment; filename="data.xlsx"')
    response.setHeader("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.write(result)


def download(self, response, delimiter=""):
    """This patch combines the data from the sender installation and the receiver
    installation to one csv / tsv.
    We assume that the form config is the same on both sides.

    The concept is to first execute the standard implementation, which sets the
    response headers and streams the local data.

    If the context is considered public, the second step is to get the data from
    the remote realms and append it to the response body (with response.write).

    This will combine the data of the all sites in a streamed http response.

    This implementation also works in local development with two plone sites,
    assuming that the receiver side does not have any realms configured.
    """

    site = getSite()
    realms = IConfig(site).getRealms()
    if not realms:
        return

    if site.REQUEST.get('is_publisher', None):
        # Prevent endless loop if sender and receiver are running on the same machine
        self._old_download(response, delimiter)
        return

    pub_state = getMultiAdapter((get_context(self), response), IPublisherContextState)
    if not pub_state.is_parent_published():
        return

    site_path = '/'.join(site.getPhysicalPath())
    context_path = '/'.join(get_context(self).getPhysicalPath())
    relative_path = os.path.relpath(context_path, site_path)
    view_path = '/'.join((relative_path, '@@actions', self.__name__, '@@data'))

    is_xlsx = getattr(self, 'DownloadFormat', 'tsv') == 'xlsx'
    is_csv = getattr(self, 'DownloadFormat', 'tsv') == 'csv'

    if is_csv and len(delimiter) == 0:
        delimiter = ','

    if not is_xlsx:
        self._old_download(response, delimiter)

    for realm in realms:
        try:
            remote_response = sendRequestToRealm(getSite().REQUEST.form.copy(),
                                                 realm,
                                                 view_path.lstrip('/') + '?is_publisher=1',
                                                 return_response=True)

        except HTTPError:
            remote_response = False  # Nothing to combine

        if remote_response and remote_response.code != 200:
            raise ValueError('Bad response from remote realm ({} {}): {!r}..'.format(
                remote_response.code,
                remote_response.msg,
                remote_response.read(100),
            ))

        try:
            if is_xlsx:
                use_title_row = getattr(self, "UseColumnNames", False)
                local_excel = self.get_saved_form_input_as_xlsx(use_title_row)
                combine_excel(response, local_excel, remote_response, use_title_row)
            else:
                response.write(remote_response.read())
        except ImportError:
            raise Exception('Was not able to combine excel, since openpyxl is missing')
        finally:
            if remote_response:
                remote_response.close()
